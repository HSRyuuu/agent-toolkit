"""xlsx → 프로파일 JSON 추출기 (결정론적).

`excel-doc-updater` 스킬의 Step 1에서 호출된다. 입력 xlsx의 구조를 읽어
LLM이 전략 추론에 사용할 수 있는 JSON 프로파일을 출력한다. 입력 파일은
read-only로 다루며 절대 수정하지 않는다.

스키마: ../references/profile_schema.md
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


ID_PATTERN = re.compile(r"^[A-Za-z]+(?:[-_][A-Za-z0-9]+){2,}$")
PREFIX_SPLIT = re.compile(r"^(?P<prefix>[A-Za-z][\w-]*?[-_])(?P<tail>\d+)$")


def detect_header_candidates(ws: Worksheet, scan_rows: int = 12) -> list[dict[str, Any]]:
    """1~scan_rows 범위에서 텍스트 비율이 높은 행을 헤더 후보로 잡는다.

    `ws.max_column`은 시트 전체 max라 헤더 행 자체가 좁은 양식에서 ratio가
    낮게 잡힌다. 그래서 행 내부의 "마지막 텍스트 셀"까지만 보고 ratio를 잰다.
    """
    candidates: list[dict[str, Any]] = []
    sheet_max_col = min(ws.max_column or 1, 50)
    for row_idx in range(1, min(scan_rows, ws.max_row or 0) + 1):
        values = [ws.cell(row=row_idx, column=c).value for c in range(1, sheet_max_col + 1)]
        last_text_idx = -1
        for i, v in enumerate(values):
            if isinstance(v, str) and v.strip():
                last_text_idx = i
        if last_text_idx < 0:
            continue
        trimmed = values[: last_text_idx + 1]
        text_cells = [v for v in trimmed if isinstance(v, str) and v.strip()]
        ratio = len(text_cells) / len(trimmed)
        if ratio >= 0.6 and len(text_cells) >= 3:
            candidates.append(
                {
                    "row": row_idx,
                    "filled_ratio": round(ratio, 3),
                    "values": [str(v) if v is not None else "" for v in trimmed],
                }
            )
    return candidates


def detect_id_columns(ws: Worksheet, scan_rows_limit: int = 500) -> list[dict[str, Any]]:
    """각 컬럼별로 ID 패턴(prefix가 같은 값)이 3회 이상 반복되는지 본다."""
    max_col = min(ws.max_column or 1, 50)
    max_row = min(ws.max_row or 1, scan_rows_limit)
    results: list[dict[str, Any]] = []
    for col_idx in range(1, max_col + 1):
        prefix_to_rows: dict[str, list[int]] = defaultdict(list)
        prefix_to_samples: dict[str, list[str]] = defaultdict(list)
        for row_idx in range(1, max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if not isinstance(val, str):
                continue
            val = val.strip()
            if not ID_PATTERN.match(val):
                continue
            m = PREFIX_SPLIT.match(val)
            if not m:
                continue
            prefix = m.group("prefix")
            prefix_to_rows[prefix].append(row_idx)
            if len(prefix_to_samples[prefix]) < 3:
                prefix_to_samples[prefix].append(val)
        for prefix, rows in prefix_to_rows.items():
            if len(rows) >= 3:
                results.append(
                    {
                        "column": get_column_letter(col_idx),
                        "prefix": prefix,
                        "match_count": len(rows),
                        "rows": rows[:10] + (["..."] if len(rows) > 10 else []),
                        "sample_values": prefix_to_samples[prefix],
                    }
                )
    return results


def style_fingerprint_for_row(ws: Worksheet, row_idx: int, max_col: int) -> dict[str, Any]:
    """헤더 다음 행 한 줄의 스타일 요약. 신규 행 복제 템플릿 후보로 사용."""
    fp: dict[str, Any] = {}
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        font = cell.font
        fill = cell.fill
        border = cell.border
        align = cell.alignment
        fp[get_column_letter(col_idx)] = {
            "font": f"{font.name or ''}/{font.size or ''}/{'bold' if font.bold else 'regular'}",
            "fill_fg": getattr(getattr(fill, "fgColor", None), "rgb", None),
            "border_left": getattr(getattr(border, "left", None), "style", None),
            "border_right": getattr(getattr(border, "right", None), "style", None),
            "align": f"{align.horizontal or ''}/{align.vertical or ''}",
            "number_format": cell.number_format,
        }
    return fp


def count_cross_sheet_formulas(ws: Worksheet) -> int:
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and "!" in v:
                count += 1
    return count


def profile_sheet(ws: Worksheet, index: int) -> dict[str, Any]:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    header_candidates = detect_header_candidates(ws)
    id_columns = detect_id_columns(ws)
    primary_header_row = header_candidates[0]["row"] if header_candidates else None
    fingerprint_row = primary_header_row + 1 if primary_header_row else 1

    column_widths: dict[str, float] = {}
    for letter, dim in (ws.column_dimensions or {}).items():
        if dim.width is not None:
            column_widths[letter] = round(float(dim.width), 2)

    row_heights: dict[str, float] = {}
    for r, dim in (ws.row_dimensions or {}).items():
        if dim.height is not None:
            row_heights[str(r)] = round(float(dim.height), 2)

    return {
        "name": ws.title,
        "index": index,
        "max_row": max_row,
        "max_col": max_col,
        "dimensions": ws.dimensions,
        "header_row_candidates": header_candidates,
        "merged_ranges": [str(rng) for rng in (ws.merged_cells.ranges or [])],
        "column_widths": column_widths,
        "row_heights": row_heights,
        "id_columns": id_columns,
        "cell_style_fingerprint": {
            f"row_{fingerprint_row}_template": style_fingerprint_for_row(
                ws, fingerprint_row, min(max_col, 20)
            )
        }
        if max_row >= fingerprint_row and max_col > 0
        else {},
        "cross_sheet_formula_count": count_cross_sheet_formulas(ws),
        "has_images": bool(getattr(ws, "_images", [])),
        "has_charts": bool(getattr(ws, "_charts", [])),
        "has_data_validation": bool(getattr(ws.data_validations, "dataValidation", [])),
        "has_conditional_formatting": bool(list(ws.conditional_formatting or [])),
    }


def cluster_sheet_prefixes(sheet_names: list[str]) -> list[dict[str, Any]]:
    """시트 이름의 prefix(끝 숫자 떼고)별 묶음."""
    buckets: dict[str, list[str]] = defaultdict(list)
    for name in sheet_names:
        m = PREFIX_SPLIT.match(name)
        if m:
            buckets[m.group("prefix")].append(name)
    clusters = []
    for prefix, names in buckets.items():
        if len(names) >= 2:
            clusters.append(
                {
                    "prefix": prefix,
                    "sheet_names": names,
                    "count": len(names),
                    "master_candidate": sorted(names)[0],
                }
            )
    return sorted(clusters, key=lambda c: -c["count"])


def duplicate_header_groups(sheet_profiles: list[dict[str, Any]]) -> list[list[str]]:
    sig_to_names: dict[tuple, list[str]] = defaultdict(list)
    for sp in sheet_profiles:
        cands = sp.get("header_row_candidates") or []
        if not cands:
            continue
        sig = tuple(cands[0]["values"])
        sig_to_names[sig].append(sp["name"])
    return [names for names in sig_to_names.values() if len(names) >= 2]


def profile_workbook(input_path: Path) -> dict[str, Any]:
    if not input_path.is_file():
        raise FileNotFoundError(f"입력 xlsx가 없습니다: {input_path}")
    wb = load_workbook(filename=str(input_path), data_only=False, read_only=False)
    sheet_profiles = [profile_sheet(wb[name], idx) for idx, name in enumerate(wb.sheetnames)]
    stat = input_path.stat()
    return {
        "input_path": str(input_path),
        "input_size_bytes": stat.st_size,
        "input_mtime": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
        "profiled_at": datetime.now(tz=timezone.utc).isoformat(),
        "sheet_count": len(sheet_profiles),
        "sheets": sheet_profiles,
        "cross_sheet_signals": {
            "sheet_name_prefix_clusters": cluster_sheet_prefixes(wb.sheetnames),
            "duplicate_header_signature_groups": duplicate_header_groups(sheet_profiles),
        },
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="xlsx 구조를 분석해 프로파일 JSON을 출력한다.")
    parser.add_argument("--input", required=True, help="대상 xlsx 경로")
    parser.add_argument("--output", help="결과 JSON 저장 경로 (생략 시 stdout)")
    args = parser.parse_args(argv)

    profile = profile_workbook(Path(args.input))
    payload = json.dumps(profile, ensure_ascii=False, indent=2, default=str)

    if args.output:
        out_path = Path(args.output)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(payload, encoding="utf-8")
        print(f"profile saved: {out_path}", file=sys.stderr)
    else:
        sys.stdout.write(payload + os.linesep)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
