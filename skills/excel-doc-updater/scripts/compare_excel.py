"""두 xlsx 파일을 비교해 markdown diff 리포트를 만든다 (결정론적).

`excel-doc-updater` 스킬의 Step 5에서 호출된다. 입력 두 파일은 read-only로
다루며 절대 수정하지 않는다.

리포트 스키마: ../references/diff_report_schema.md
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def cell_value(ws: Worksheet, row: int, col: int) -> Any:
    return ws.cell(row=row, column=col).value


def collect_used_cells(ws: Worksheet) -> dict[tuple[int, int], Any]:
    """시트의 비어있지 않은 모든 셀(또는 max_row x max_col 범위)의 값."""
    cells: dict[tuple[int, int], Any] = {}
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                cells[(r, c)] = v
    return cells


def diff_sheet_cells(
    before: Worksheet, after: Worksheet, sample_limit: int
) -> dict[str, Any]:
    bcells = collect_used_cells(before)
    acells = collect_used_cells(after)

    keys = set(bcells.keys()) | set(acells.keys())
    changes: list[dict[str, Any]] = []
    for key in sorted(keys):
        bv = bcells.get(key)
        av = acells.get(key)
        if bv != av:
            changes.append(
                {
                    "cell": f"{get_column_letter(key[1])}{key[0]}",
                    "before": bv,
                    "after": av,
                }
            )

    return {
        "changed_cell_count": len(changes),
        "samples": changes[:sample_limit],
        "before_max_row": before.max_row or 0,
        "after_max_row": after.max_row or 0,
        "before_max_col": before.max_column or 0,
        "after_max_col": after.max_column or 0,
    }


def render_report(
    before_path: Path,
    after_path: Path,
    protected: list[str],
    sample_cells: int,
) -> str:
    wb_before = load_workbook(filename=str(before_path), data_only=False, read_only=False)
    wb_after = load_workbook(filename=str(after_path), data_only=False, read_only=False)

    before_names = list(wb_before.sheetnames)
    after_names = list(wb_after.sheetnames)
    before_set = set(before_names)
    after_set = set(after_names)

    added_sheets = [n for n in after_names if n not in before_set]
    removed_sheets = [n for n in before_names if n not in after_set]
    common_sheets = [n for n in after_names if n in before_set]

    sheet_diffs: dict[str, dict[str, Any]] = {}
    total_changed_cells = 0
    protected_changes: list[dict[str, Any]] = []

    for name in common_sheets:
        d = diff_sheet_cells(wb_before[name], wb_after[name], sample_cells)
        if d["changed_cell_count"] > 0:
            sheet_diffs[name] = d
            total_changed_cells += d["changed_cell_count"]
            if name in protected:
                for s in d["samples"]:
                    protected_changes.append({"sheet": name, **s})

    rows_added_total = 0
    rows_removed_total = 0
    for name, d in sheet_diffs.items():
        delta = d["after_max_row"] - d["before_max_row"]
        if delta > 0:
            rows_added_total += delta
        elif delta < 0:
            rows_removed_total += -delta

    lines: list[str] = []
    lines.append("# 엑셀 변경 비교 리포트")
    lines.append("")
    lines.append(f"- before: `{before_path}`")
    lines.append(f"- after:  `{after_path}`")
    lines.append(f"- generated_at: {datetime.now(tz=timezone.utc).isoformat()}")
    lines.append(f"- protected_sheets: {protected if protected else '(없음)'}")
    lines.append("")
    lines.append("## 요약")
    lines.append("")
    lines.append("| 항목 | 값 |")
    lines.append("|---|---|")
    lines.append(f"| 시트 추가 | {len(added_sheets)} |")
    lines.append(f"| 시트 삭제 | {len(removed_sheets)} |")
    lines.append(f"| 행 추가 (총 max_row 증가분) | {rows_added_total} |")
    lines.append(f"| 행 삭제 (총 max_row 감소분) | {rows_removed_total} |")
    lines.append(f"| 변경 셀 (총) | {total_changed_cells} |")
    protected_marker = (
        f"**{len(protected_changes)}**" if protected_changes else "0"
    )
    lines.append(f"| 보호 영역 변경 | {protected_marker} |")
    lines.append("")

    if protected_changes:
        lines.append("## ⚠️ 보호 영역 경고")
        lines.append("")
        lines.append("| sheet | cell | before | after |")
        lines.append("|---|---|---|---|")
        for ch in protected_changes[: sample_cells * 4]:
            lines.append(
                f"| {ch['sheet']} | {ch['cell']} | {ch['before']!r} | {ch['after']!r} |"
            )
        lines.append("")
        lines.append(
            "→ 의도하지 않은 변경이라면 출력 파일을 폐기하는 것을 권장합니다."
        )
        lines.append("")

    lines.append("## 시트 단위 변경")
    lines.append("")
    if added_sheets:
        lines.append("### 신규 시트")
        for name in added_sheets:
            ws = wb_after[name]
            lines.append(
                f"- `{name}` (rows={ws.max_row or 0}, cols={ws.max_column or 0})"
            )
        lines.append("")
    if removed_sheets:
        lines.append("### 삭제된 시트")
        for name in removed_sheets:
            lines.append(f"- `{name}`")
        lines.append("")

    if not sheet_diffs:
        lines.append("(공통 시트 변경 없음)")
        lines.append("")
    else:
        for name, d in sheet_diffs.items():
            tag = " (보호 영역)" if name in protected else ""
            lines.append(f"### {name}{tag}")
            lines.append("")
            lines.append(
                f"- 변경 셀: {d['changed_cell_count']}개, "
                f"max_row {d['before_max_row']}→{d['after_max_row']}, "
                f"max_col {d['before_max_col']}→{d['after_max_col']}"
            )
            if d["samples"]:
                lines.append("")
                lines.append("| cell | before | after |")
                lines.append("|---|---|---|")
                for s in d["samples"]:
                    lines.append(
                        f"| {s['cell']} | {s['before']!r} | {s['after']!r} |"
                    )
            lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="두 xlsx를 비교해 markdown diff 리포트를 만든다."
    )
    parser.add_argument("--before", required=True, help="원본 xlsx")
    parser.add_argument("--after", required=True, help="결과 xlsx")
    parser.add_argument(
        "--protected",
        default="",
        help="보호 시트 이름 (콤마 구분). 변경이 감지되면 WARNING 섹션으로 표시.",
    )
    parser.add_argument(
        "--sample-cells",
        type=int,
        default=5,
        help="시트별 변경 샘플 표시 최대 개수 (기본 5)",
    )
    parser.add_argument("--output", help="리포트 저장 경로 (생략 시 stdout)")
    args = parser.parse_args(argv)

    protected = [s.strip() for s in args.protected.split(",") if s.strip()]
    report = render_report(
        Path(args.before), Path(args.after), protected, args.sample_cells
    )

    if args.output:
        out_path = Path(args.output)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(report, encoding="utf-8")
        print(f"report saved: {out_path}", file=sys.stderr)
    else:
        sys.stdout.write(report)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
