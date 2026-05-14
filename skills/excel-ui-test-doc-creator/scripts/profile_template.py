"""사용자 템플릿 .xlsx → 헤더·매핑 프로파일 JSON 산출기 (결정론적).

`excel-ui-test-doc-creator` 스킬의 작성 전 단계에서 호출된다.
LLM이 템플릿을 직접 읽고 매핑을 추론하지 않게 하기 위함. 본 스크립트가
헤더 행을 탐지하고 내장 alias 사전으로 표준 키 매칭 후보를 산출한다.
결과 JSON을 그대로 사용자에게 보여주고 승인을 받은 뒤에야
`create_test_doc.py`를 호출한다.

스키마는 README/SKILL.md "프로파일 JSON" 절 참고.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# create_test_doc.py와 정확히 동일한 alias 사전을 사용한다. 둘이 어긋나면
# profile 단계가 "매칭됨"이라 보고했는데 실제 작성에선 누락되는 사고가 난다.
STANDARD_KEYS = ["id", "name", "data", "steps", "expected", "url", "actual", "status", "note", "tester", "date"]

_ALIAS_GROUPS: dict[str, list[str]] = {
    "id": ["단위테스트id", "테스트id", "tcid", "tcno", "no", "번호", "id"],
    "name": ["단위테스트명", "테스트명", "시나리오명", "테스트케이스", "테스트케이스명", "name", "case", "casename"],
    "data": ["테스트데이터", "입력데이터", "입력값", "input", "data", "inputdata"],
    "steps": ["테스트절차", "절차", "수행절차", "수행단계", "테스트단계", "steps", "procedure", "step"],
    "expected": ["기대결과", "예상결과", "expected", "예상", "expectedresult"],
    "url": ["화면url", "url", "화면", "screen", "screenurl", "page"],
    "actual": ["실제결과", "수행결과", "결과", "actual", "result", "actualresult"],
    "status": ["상태", "합격여부", "passfail", "결과상태", "status", "verdict"],
    "note": ["비고", "메모", "note", "remarks", "comment"],
    "tester": ["테스터", "수행자", "담당자", "tester"],
    "date": ["수행일", "테스트일", "date", "testedat", "testdate"],
}
HEADER_ALIASES: dict[str, str] = {}
for _key, _aliases in _ALIAS_GROUPS.items():
    for _a in _aliases:
        HEADER_ALIASES[_a] = _key


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[\s:_\-/]+", "", text)
    return text


def detect_header_row(ws: Worksheet, scan_rows: int = 12) -> tuple[int, list[str]]:
    max_col = min(ws.max_column or 1, 50)
    best_row = 1
    best_score = -1.0
    best_values: list[str] = []
    for row_idx in range(1, min(scan_rows, ws.max_row or 0) + 1):
        values = [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
        last_text_idx = -1
        for i, v in enumerate(values):
            if isinstance(v, str) and v.strip():
                last_text_idx = i
        if last_text_idx < 2:
            continue
        trimmed = values[: last_text_idx + 1]
        text_cells = [v for v in trimmed if isinstance(v, str) and v.strip()]
        ratio = len(text_cells) / len(trimmed)
        score = ratio * len(text_cells)
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_values = [str(v) if v is not None else "" for v in trimmed]
    return best_row, best_values


def style_sample(ws: Worksheet, row: int, col: int) -> dict[str, Any]:
    cell = ws.cell(row=row, column=col)
    font_name = cell.font.name if cell.font else None
    font_size = cell.font.size if cell.font else None
    bold = bool(cell.font.bold) if cell.font else False
    fill_rgb = None
    if cell.fill and cell.fill.start_color and cell.fill.fill_type:
        fill_rgb = cell.fill.start_color.rgb
    align_h = cell.alignment.horizontal if cell.alignment else None
    align_v = cell.alignment.vertical if cell.alignment else None
    wrap = bool(cell.alignment.wrap_text) if cell.alignment else False
    return {
        "font_name": font_name,
        "font_size": font_size,
        "bold": bold,
        "fill_rgb": fill_rgb,
        "alignment_horizontal": align_h,
        "alignment_vertical": align_v,
        "wrap_text": wrap,
    }


def detect_merged_ranges_above_header(ws: Worksheet, header_row: int) -> list[str]:
    out: list[str] = []
    for rng in ws.merged_cells.ranges:
        if rng.max_row < header_row:
            out.append(str(rng))
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description="템플릿 xlsx 헤더·매핑 프로파일러")
    parser.add_argument("--template", required=True, help="사용자 양식 .xlsx 경로")
    parser.add_argument("--sheet", help="시트 이름 (미지정 시 active)")
    parser.add_argument("--output", help="프로파일 JSON 저장 경로 (미지정 시 stdout)")
    args = parser.parse_args()

    template_path = Path(args.template)
    if not template_path.exists():
        print(f"[ERROR] 템플릿이 없습니다: {template_path}", file=sys.stderr)
        return 2
    if template_path.suffix.lower() != ".xlsx":
        print(
            f"[ERROR] 템플릿 확장자가 .xlsx가 아닙니다: {template_path.suffix}. "
            f".xlsx로 저장 후 다시 실행해주세요.",
            file=sys.stderr,
        )
        return 2

    wb = load_workbook(template_path, data_only=True)
    if args.sheet:
        if args.sheet not in wb.sheetnames:
            print(f"[ERROR] 시트 '{args.sheet}' 를 찾을 수 없습니다. 후보: {wb.sheetnames}", file=sys.stderr)
            return 2
        ws = wb[args.sheet]
    else:
        ws = wb.active

    header_row, header_values = detect_header_row(ws)
    if not header_values:
        print(
            "[ERROR] 헤더 행을 찾지 못했습니다. 첫 12행 안에 텍스트 컬럼명이 3개 이상 있는 행이 필요합니다.",
            file=sys.stderr,
        )
        return 3

    headers: list[dict[str, Any]] = []
    matched_std_keys: set[str] = set()
    unmapped: list[dict[str, Any]] = []
    duplicate_matches: dict[str, list[int]] = {}

    for idx, raw in enumerate(header_values, start=1):
        norm = normalize_header(raw)
        std_key = HEADER_ALIASES.get(norm)
        record = {
            "col": idx,
            "col_letter": get_column_letter(idx),
            "raw": raw,
            "normalized": norm,
            "matched_standard_key": std_key,
            "confidence": "alias" if std_key else None,
        }
        headers.append(record)
        if std_key:
            duplicate_matches.setdefault(std_key, []).append(idx)
            matched_std_keys.add(std_key)
        else:
            unmapped.append({"col": idx, "col_letter": get_column_letter(idx), "raw": raw})

    duplicates = {k: v for k, v in duplicate_matches.items() if len(v) > 1}

    # 다음 빈 데이터 행 (작성 시작 행 후보)
    sample_data_row = header_row + 1
    first_data_styles = []
    if (ws.max_row or 0) >= sample_data_row:
        for h in headers:
            first_data_styles.append({"col": h["col"], **style_sample(ws, sample_data_row, h["col"])})

    profile = {
        "template_path": str(template_path),
        "sheet_title": ws.title,
        "header_row": header_row,
        "header_count": len(headers),
        "headers": headers,
        "unmapped_headers": unmapped,
        "duplicate_standard_keys": duplicates,
        "standard_keys_matched": sorted(matched_std_keys),
        "standard_keys_not_in_template": [k for k in STANDARD_KEYS if k not in matched_std_keys],
        "merged_ranges_above_header": detect_merged_ranges_above_header(ws, header_row),
        "max_existing_data_row": max(ws.max_row or 0, header_row),
        "first_data_row_sample_styles": {"row": sample_data_row, "columns": first_data_styles},
    }

    payload = json.dumps(profile, ensure_ascii=False, indent=2)
    if args.output:
        out_path = Path(args.output)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(payload, encoding="utf-8")
        print(f"[OK] 프로파일 저장: {out_path}")
    else:
        print(payload)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
