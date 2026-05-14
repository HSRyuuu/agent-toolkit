"""작성된 단위테스트 산출물 .xlsx 결정론적 검증 리포터.

`excel-ui-test-doc-creator` 스킬의 작성 후 단계에서 호출된다.
입력 spec(JSON)과 출력 xlsx를 다시 읽어, 각 spec 한 건이 어느 행·어느 셀에
어떤 값으로 들어갔는지를 결정론적으로 표로 출력한다. LLM이 "잘 들어간 것
같아요" 식의 spot check로 완료 보고하는 것을 막기 위함.

CLI:
    python3 verify_test_doc.py --input specs.json --output out.xlsx [--mapping map.json]
    # 템플릿 모드면 --mapping 필요. 없으면 내장 alias로 자동 추정.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


STANDARD_KEYS = ["id", "name", "data", "steps", "expected", "url", "actual", "status", "note", "tester", "date"]

_ALIAS_GROUPS: dict[str, list[str]] = {
    "id": ["단위테스트id", "테스트id", "tcid", "tcno", "no", "번호", "id"],
    "name": ["단위테스트명", "테스트명", "시나리오명", "테스트케이스", "테스트케이스명", "name", "case", "casename"],
    "data": ["테스트데이터", "입력데이터", "입력값", "input", "data", "inputdata"],
    "steps": ["테스트절차", "절차", "수행절차", "수행단계", "테스트단계", "steps", "procedure", "step"],
    "expected": ["기대결과", "예상결과", "expected", "예상", "expectedresult"],
    "url": ["화면url", "url", "화면", "screen", "screenurl", "page"],
    "actual": ["실제결과", "수행결과", "결과", "actual", "result", "actualresult"],
    "status": ["상태", "합격여부", "passfail", "결과상태", "status", "verdict"],
    "note": ["비고", "메모", "note", "remarks", "comment"],
    "tester": ["테스터", "수행자", "담당자", "tester"],
    "date": ["수행일", "테스트일", "date", "testedat", "testdate"],
}
HEADER_ALIASES: dict[str, str] = {}
for _key, _aliases in _ALIAS_GROUPS.items():
    for _a in _aliases:
        HEADER_ALIASES[_a] = _key


DEFAULT_HEADERS_KO = {
    "id": "단위테스트 ID",
    "name": "단위테스트명",
    "data": "테스트 데이터",
    "steps": "테스트 절차",
    "expected": "기대 결과",
    "url": "화면 URL",
    "actual": "실제 결과",
}

STATUS_FILL: dict[str, str] = {
    "pass": "E2EFDA",
    "fail": "FCE4D6",
    "block": "FFF2CC",
    "na": "F2F2F2",
}


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[\s:_\-/]+", "", text)
    return text


def join_steps(steps: Any) -> str:
    if steps is None:
        return ""
    if isinstance(steps, str):
        return steps.strip()
    if isinstance(steps, list):
        parts = [str(s).strip() for s in steps if s is not None and str(s).strip()]
        return "\n".join(parts)
    return str(steps)


def value_for_key(test: dict[str, Any], key: str) -> str:
    raw = test.get(key, "")
    if key == "steps":
        return join_steps(raw)
    if raw is None:
        return ""
    if isinstance(raw, (list, dict)):
        return json.dumps(raw, ensure_ascii=False)
    return str(raw)


def detect_header_row(ws: Worksheet, scan_rows: int = 12) -> tuple[int, list[str]]:
    max_col = min(ws.max_column or 1, 50)
    best_row = 1
    best_score = -1.0
    best_values: list[str] = []
    for row_idx in range(1, min(scan_rows, ws.max_row or 0) + 1):
        values = [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
        last_text_idx = -1
        for i, v in enumerate(values):
            if isinstance(v, str) and v.strip():
                last_text_idx = i
        if last_text_idx < 2:
            continue
        trimmed = values[: last_text_idx + 1]
        text_cells = [v for v in trimmed if isinstance(v, str) and v.strip()]
        ratio = len(text_cells) / len(trimmed)
        score = ratio * len(text_cells)
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_values = [str(v) if v is not None else "" for v in trimmed]
    return best_row, best_values


def load_specs(input_path: Path) -> list[dict[str, Any]]:
    with input_path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        return data.get("tests") or data.get("cases") or data.get("items") or []
    return []


def truncate(text: str, width: int) -> str:
    text = text.replace("\n", " ⏎ ")
    if len(text) <= width:
        return text
    return text[: width - 1] + "…"


def main() -> int:
    parser = argparse.ArgumentParser(description="작성된 테스트 산출물 xlsx 결정론적 검증")
    parser.add_argument("--input", required=True, help="원본 spec JSON")
    parser.add_argument("--output", required=True, help="검증 대상 .xlsx")
    parser.add_argument("--mapping", help="헤더 → 표준 키 매핑 JSON (없으면 자동 추정)")
    parser.add_argument("--sheet", help="시트 이름 (미지정 시 active)")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    if not input_path.exists():
        print(f"[ERROR] spec JSON이 없습니다: {input_path}", file=sys.stderr)
        return 2
    if not output_path.exists():
        print(f"[ERROR] 검증 대상 xlsx가 없습니다: {output_path}", file=sys.stderr)
        return 2

    specs = load_specs(input_path)
    wb = load_workbook(output_path)
    ws = wb[args.sheet] if args.sheet else wb.active

    header_row, header_values = detect_header_row(ws)
    if not header_values:
        print("[ERROR] 헤더 행을 찾지 못했습니다.", file=sys.stderr)
        return 3

    # 매핑 결정
    user_mapping: dict[str, str] = {}
    if args.mapping:
        with Path(args.mapping).open("r", encoding="utf-8") as f:
            user_mapping = json.load(f)
        if not isinstance(user_mapping, dict):
            print("[ERROR] mapping JSON은 dict여야 합니다.", file=sys.stderr)
            return 2

    col_to_key: dict[int, str] = {}
    for idx, raw in enumerate(header_values, start=1):
        std = user_mapping.get(raw) or user_mapping.get(str(raw).strip())
        if not std:
            std = HEADER_ALIASES.get(normalize_header(raw))
        if std:
            col_to_key[idx] = std

    # 데이터 행 범위: header_row + 1 ~ header_row + len(specs)
    data_start = header_row + 1
    data_end = data_start + len(specs) - 1
    actual_max_row = ws.max_row or header_row
    row_count_xlsx = max(0, actual_max_row - header_row) if actual_max_row > header_row else 0

    # ------------- 출력 -------------
    print("=" * 72)
    print(" 단위테스트 산출물 검증 리포트")
    print("=" * 72)
    print(f" 파일       : {output_path}")
    print(f" 시트       : {ws.title}")
    print(f" 헤더 행    : {header_row}  →  {' | '.join(header_values)}")
    print(f" 데이터 행  : r{data_start}~r{data_end}  (spec {len(specs)}건)")
    if row_count_xlsx != len(specs):
        print(f" [!] xlsx 실제 데이터 행 수={row_count_xlsx} ≠ spec 건수={len(specs)}")
    else:
        print(f" 행 수 일치 : ✓ {len(specs)}건")

    # 매핑 표
    print("\n[헤더 ↔ 표준 키 매핑]")
    for idx, raw in enumerate(header_values, start=1):
        std = col_to_key.get(idx, "—  (미매핑: 빈 값으로 둠)")
        print(f"  {get_column_letter(idx):>2}({idx})  '{raw}'  →  {std}")

    # spec 키 중 템플릿에 없는 것
    spec_keys: set[str] = set()
    for s in specs:
        for k in s.keys():
            if k in STANDARD_KEYS:
                spec_keys.add(k)
    cols_keys = set(col_to_key.values())
    dropped = sorted(spec_keys - cols_keys)
    if dropped:
        print(f"\n[!] spec에 있으나 대상 양식에 컬럼이 없어 누락된 키: {dropped}")

    # 행별 비교
    print("\n[행별 결정론적 비교]")
    mismatches: list[str] = []
    status_counts: dict[str, int] = {}
    fill_mismatches: list[str] = []

    for i, spec in enumerate(specs):
        row = data_start + i
        spec_id = spec.get("id", "")
        spec_status = (spec.get("status") or "").strip().lower()
        status_counts[spec_status or "empty"] = status_counts.get(spec_status or "empty", 0) + 1

        per_cell = []
        for col, key in sorted(col_to_key.items()):
            expected = value_for_key(spec, key)
            actual_raw = ws.cell(row=row, column=col).value
            actual = "" if actual_raw is None else str(actual_raw)
            ok = (expected.strip() == actual.strip()) or (expected == "" and actual == "")
            mark = "✓" if ok else "✗"
            if not ok:
                mismatches.append(
                    f"r{row} {get_column_letter(col)} ({key}): expected={truncate(expected, 40)!r}  actual={truncate(actual, 40)!r}"
                )
            per_cell.append(f"{mark}{key}")
        print(f"  r{row}  {spec_id:<10}  status={spec_status or '∅':<6}  | " + " ".join(per_cell))

        # 색상 검증
        expected_fill = STATUS_FILL.get(spec_status)
        first_cell_fill = ws.cell(row=row, column=min(col_to_key) if col_to_key else 1).fill
        actual_fill_rgb = None
        if first_cell_fill and first_cell_fill.start_color and first_cell_fill.fill_type:
            rgb = first_cell_fill.start_color.rgb or ""
            if isinstance(rgb, str) and len(rgb) >= 6 and rgb[-6:] != "000000":
                actual_fill_rgb = rgb[-6:].upper()
        expected_norm = expected_fill.upper() if expected_fill else None
        if expected_norm != actual_fill_rgb:
            fill_mismatches.append(
                f"r{row} status={spec_status or '∅'}  expected_fill={expected_norm}  actual_fill={actual_fill_rgb}"
            )

    # 빈 셀 (spec에 값 없어서 의도적으로 비운 셀) 리스트
    print("\n[의도적으로 빈 셀 (spec에 값 없음)]")
    empty_intentional: list[str] = []
    for i, spec in enumerate(specs):
        row = data_start + i
        for col, key in sorted(col_to_key.items()):
            expected = value_for_key(spec, key)
            actual = ws.cell(row=row, column=col).value
            if expected == "" and (actual is None or str(actual) == ""):
                empty_intentional.append(f"r{row} {key}")
    if empty_intentional:
        for line in empty_intentional[:20]:
            print(f"  - {line}")
        if len(empty_intentional) > 20:
            print(f"  …외 {len(empty_intentional) - 20}건")
    else:
        print("  (없음)")

    print("\n[상태별 카운트]")
    for k, v in sorted(status_counts.items(), key=lambda x: x[0]):
        print(f"  {k:<6} : {v}")

    print("\n[행 색상 적용 검증]")
    if fill_mismatches:
        for line in fill_mismatches:
            print(f"  ✗ {line}")
    else:
        print("  ✓ 모든 행의 status별 배경색이 일치")

    # ------------- 결과 코드 -------------
    print("\n" + "=" * 72)
    if mismatches:
        print(f" 셀 값 불일치 {len(mismatches)}건:")
        for line in mismatches[:30]:
            print(f"  ✗ {line}")
        if len(mismatches) > 30:
            print(f"  …외 {len(mismatches) - 30}건")
        print(" 검증 실패")
        print("=" * 72)
        return 1

    if fill_mismatches:
        print(" 셀 값은 일치하나 일부 행 색상 불일치 (위 참조)")
        print("=" * 72)
        return 1

    print(" 검증 통과 ✓")
    print("=" * 72)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
