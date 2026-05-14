"""테스트 spec JSON → 단위테스트 산출물 .xlsx 생성기.

`excel-ui-test-doc-creator` 스킬의 Step 4에서 호출된다.
사용자 템플릿이 있으면 그 양식의 헤더·스타일을 보존하며 데이터 행만 채우고,
없으면 기본 양식으로 신규 통합문서를 만든다.

CLI 예:
    python3 create_test_doc.py --input specs.json --output out.xlsx
    python3 create_test_doc.py --input specs.json --template form.xlsx --mapping map.json --output out.xlsx

mapping.json 형식 (선택):
    {"단위테스트 ID": "id", "테스트 절차": "steps", ...}
    헤더 텍스트 → 표준 키. 미제공 시 내장 alias 사전으로 자동 매칭.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


STANDARD_KEYS = ["id", "name", "data", "steps", "expected", "url", "actual", "status", "note", "tester", "date"]

DEFAULT_HEADERS_KO = {
    "id": "단위테스트 ID",
    "name": "단위테스트명",
    "data": "테스트 데이터",
    "steps": "테스트 절차",
    "expected": "기대 결과",
    "url": "화면 URL",
    "actual": "실제 결과",
}

# 헤더 텍스트 → 표준 키. 비교 시 normalize_header()로 정규화.
HEADER_ALIASES: dict[str, str] = {}
_ALIAS_GROUPS = {
    "id": ["단위테스트id", "테스트id", "tcid", "tcno", "no", "번호", "id"],
    "name": ["단위테스트명", "테스트명", "시나리오명", "테스트케이스", "테스트케이스명", "name", "case", "casename"],
    "data": ["테스트데이터", "입력데이터", "입력값", "input", "data", "inputdata"],
    "steps": ["테스트절차", "절차", "수행절차", "수행단계", "테스트단계", "steps", "procedure", "step"],
    "expected": ["기대결과", "예상결과", "expected", "예상", "expectedresult"],
    "url": ["화면url", "url", "화면", "screen", "screenurl", "page"],
    "actual": ["실제결과", "수행결과", "결과", "actual", "result", "actualresult"],
    "status": ["상태", "합격여부", "passfail", "결과상태", "status", "verdict"],
    "note": ["비고", "메모", "note", "remarks", "comment"],
    "tester": ["테스터", "수행자", "담당자", "tester"],
    "date": ["수행일", "테스트일", "date", "testedat", "testdate"],
}
for _key, _aliases in _ALIAS_GROUPS.items():
    for _a in _aliases:
        HEADER_ALIASES[_a] = _key

STATUS_FILL: dict[str, str] = {
    "pass": "E2EFDA",
    "fail": "FCE4D6",
    "block": "FFF2CC",
    "na": "F2F2F2",
}

HEADER_FILL_HEX = "D9E1F2"
LONG_TEXT_THRESHOLD = 40


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[\s:_\-/]+", "", text)
    return text


def auto_map_headers(headers: list[str]) -> dict[int, str]:
    """헤더 텍스트 리스트 → {컬럼 인덱스(1-based): 표준 키} 매핑.

    매칭되지 않은 컬럼은 결과에 포함하지 않는다 (호출 측에서 사용자 확인 분기).
    """
    mapping: dict[int, str] = {}
    for idx, header in enumerate(headers, start=1):
        key = HEADER_ALIASES.get(normalize_header(header))
        if key:
            mapping[idx] = key
    return mapping


def detect_header_row(ws: Worksheet, scan_rows: int = 12) -> tuple[int, list[str]]:
    """텍스트 비율이 가장 높은 상단 행을 헤더로 본다.

    예외: 1행이 병합된 제목/타이틀이고 2~3행에 실제 컬럼명이 오는 양식도 있어
    text ratio가 가장 높은 행을 채택한다.
    """
    max_col = min(ws.max_column or 1, 50)
    best_row = 1
    best_score = -1.0
    best_values: list[str] = []
    for row_idx in range(1, min(scan_rows, ws.max_row or 0) + 1):
        values = [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
        last_text_idx = -1
        for i, v in enumerate(values):
            if isinstance(v, str) and v.strip():
                last_text_idx = i
        if last_text_idx < 2:
            continue
        trimmed = values[: last_text_idx + 1]
        text_cells = [v for v in trimmed if isinstance(v, str) and v.strip()]
        ratio = len(text_cells) / len(trimmed)
        score = ratio * len(text_cells)
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_values = [str(v) if v is not None else "" for v in trimmed]
    return best_row, best_values


def join_steps(steps: Any) -> str:
    if steps is None:
        return ""
    if isinstance(steps, str):
        return steps.strip()
    if isinstance(steps, list):
        parts = []
        for s in steps:
            if s is None:
                continue
            parts.append(str(s).strip())
        return "\n".join(p for p in parts if p)
    return str(steps)


def value_for_key(test: dict[str, Any], key: str) -> str:
    """spec dict에서 표준 키에 해당하는 셀 텍스트를 뽑는다."""
    raw = test.get(key, "")
    if key == "steps":
        return join_steps(raw)
    if raw is None:
        return ""
    if isinstance(raw, (list, dict)):
        return json.dumps(raw, ensure_ascii=False)
    return str(raw)


def apply_body_alignment(cell, has_long_text: bool) -> None:
    cell.alignment = Alignment(
        wrap_text=True if has_long_text else False,
        vertical="top",
    )


def fill_for_status(status: str) -> PatternFill | None:
    key = (status or "").strip().lower()
    hex_color = STATUS_FILL.get(key)
    if not hex_color:
        return None
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def find_first_empty_row(ws: Worksheet, header_row: int, max_col: int) -> int:
    """header_row 이후 첫 빈 데이터 행을 찾는다 (없으면 header_row + 1)."""
    for r in range(header_row + 1, (ws.max_row or header_row) + 2):
        empty = True
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                empty = False
                break
        if empty:
            return r
    return header_row + 1


def write_into_template(
    template_path: Path,
    tests: list[dict[str, Any]],
    output_path: Path,
    user_mapping: dict[str, str] | None,
) -> dict[str, Any]:
    wb = load_workbook(template_path)
    ws = wb.active

    header_row, headers = detect_header_row(ws)
    if not headers:
        raise RuntimeError(
            f"템플릿 '{template_path}'에서 헤더 행을 찾지 못했습니다. 첫 12행 안에 컬럼명이 있어야 합니다."
        )

    if user_mapping:
        col_map: dict[int, str] = {}
        for idx, header in enumerate(headers, start=1):
            std = user_mapping.get(header) or user_mapping.get(header.strip())
            if std:
                col_map[idx] = std
    else:
        col_map = auto_map_headers(headers)

    if not col_map:
        raise RuntimeError(
            "템플릿 헤더 중 어느 컬럼도 표준 키로 매핑되지 않았습니다. "
            "--mapping 옵션으로 명시적 매핑을 전달해주세요."
        )

    start_row = find_first_empty_row(ws, header_row, len(headers))

    # 헤더 행 바로 위(혹은 첫 데이터 행) 스타일을 템플릿으로 삼는다.
    template_row = header_row + 1
    has_template_row_styles = (ws.max_row or 0) >= template_row

    status_counts: dict[str, int] = {}
    for i, test in enumerate(tests):
        row = start_row + i
        for col_idx, std_key in col_map.items():
            text = value_for_key(test, std_key)
            cell = ws.cell(row=row, column=col_idx)
            cell.value = text if text != "" else None
            if has_template_row_styles:
                src = ws.cell(row=template_row, column=col_idx)
                if src.font:
                    cell.font = copy(src.font)
                if src.alignment:
                    cell.alignment = copy(src.alignment)
                if src.border:
                    cell.border = copy(src.border)
                if src.number_format:
                    cell.number_format = src.number_format
            if text and (len(text) >= LONG_TEXT_THRESHOLD or "\n" in text):
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical=(cell.alignment.vertical if cell.alignment else "top") or "top",
                    horizontal=cell.alignment.horizontal if cell.alignment else None,
                )

        status = (test.get("status") or "").strip().lower()
        status_counts[status or "empty"] = status_counts.get(status or "empty", 0) + 1
        fill = fill_for_status(status)
        if fill:
            for col_idx in col_map:
                ws.cell(row=row, column=col_idx).fill = fill

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {
        "mode": "template",
        "template": str(template_path),
        "output": str(output_path),
        "header_row": header_row,
        "start_row": start_row,
        "columns_mapped": {int(k): v for k, v in col_map.items()},
        "status_counts": status_counts,
        "rows_written": len(tests),
    }


def write_default(tests: list[dict[str, Any]], output_path: Path, title: str | None) -> dict[str, Any]:
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "단위테스트")[:31] or "단위테스트"

    keys = list(DEFAULT_HEADERS_KO.keys())
    headers = [DEFAULT_HEADERS_KO[k] for k in keys]

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color=HEADER_FILL_HEX, end_color=HEADER_FILL_HEX, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    column_widths = {"id": 14, "name": 28, "data": 26, "steps": 40, "expected": 28, "url": 32, "actual": 28}
    for col_idx, key in enumerate(keys, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = column_widths.get(key, 20)

    status_counts: dict[str, int] = {}
    for row_offset, test in enumerate(tests, start=2):
        for col_idx, key in enumerate(keys, start=1):
            text = value_for_key(test, key)
            cell = ws.cell(row=row_offset, column=col_idx, value=text if text != "" else None)
            has_long = bool(text) and (len(text) >= LONG_TEXT_THRESHOLD or "\n" in text)
            apply_body_alignment(cell, has_long_text=has_long or key == "steps")

        status = (test.get("status") or "").strip().lower()
        status_counts[status or "empty"] = status_counts.get(status or "empty", 0) + 1
        fill = fill_for_status(status)
        if fill:
            for col_idx in range(1, len(keys) + 1):
                ws.cell(row=row_offset, column=col_idx).fill = fill

    ws.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {
        "mode": "default",
        "output": str(output_path),
        "status_counts": status_counts,
        "rows_written": len(tests),
        "columns": headers,
    }


def load_specs(input_path: Path) -> tuple[list[dict[str, Any]], str | None]:
    with input_path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, list):
        return data, None
    if isinstance(data, dict):
        tests = data.get("tests") or data.get("cases") or data.get("items")
        if not isinstance(tests, list):
            raise RuntimeError(
                "입력 JSON의 최상위가 dict인 경우 'tests' (또는 'cases'/'items') 배열 키가 필요합니다."
            )
        return tests, data.get("title")
    raise RuntimeError("입력 JSON은 배열 또는 'tests' 키를 가진 dict여야 합니다.")


def main() -> int:
    parser = argparse.ArgumentParser(description="단위테스트 산출물 .xlsx 생성기")
    parser.add_argument("--input", required=True, help="테스트 spec JSON 경로")
    parser.add_argument("--template", help="사용자 양식 .xlsx (선택)")
    parser.add_argument("--mapping", help="헤더 → 표준 키 매핑 JSON (선택)")
    parser.add_argument("--output", default="./test-results.xlsx", help="출력 .xlsx 경로")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        print(f"[ERROR] 입력 JSON이 없습니다: {input_path}", file=sys.stderr)
        return 2

    tests, title = load_specs(input_path)

    user_mapping: dict[str, str] | None = None
    if args.mapping:
        mapping_path = Path(args.mapping)
        if not mapping_path.exists():
            print(f"[ERROR] mapping JSON이 없습니다: {mapping_path}", file=sys.stderr)
            return 2
        with mapping_path.open("r", encoding="utf-8") as f:
            user_mapping = json.load(f)
        if not isinstance(user_mapping, dict):
            print("[ERROR] mapping JSON은 {헤더: 표준키} dict여야 합니다.", file=sys.stderr)
            return 2

    if args.template:
        template_path = Path(args.template)
        if not template_path.exists():
            print(f"[ERROR] 템플릿이 없습니다: {template_path}", file=sys.stderr)
            return 2
        if template_path.suffix.lower() != ".xlsx":
            print(
                f"[ERROR] 템플릿 확장자가 .xlsx가 아닙니다: {template_path.suffix}. "
                f".xlsx로 저장 후 다시 실행해주세요.",
                file=sys.stderr,
            )
            return 2
        result = write_into_template(template_path, tests, output_path, user_mapping)
    else:
        result = write_default(tests, output_path, title)

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
